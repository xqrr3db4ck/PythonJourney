import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from math import ceil
import os

df = pd.read_excel('1a1.xlsx')
df.drop(df.columns[[1, 2, 3, 4, 8, 17, 18, 19]], axis=1, inplace=True)
df.iloc[:, 0] = df.iloc[:, 0].str[5:]
df.iloc[:, 1] = df.iloc[:, 1].str[5:]
df.iloc[:, 3] = df.iloc[:, 3].str[6:]
df['COD_PUB'] = df['COD_PUB'].str.lstrip('0') + '_'
df['TITULO'] = df['TITULO'].str.slice(stop=30)
df['TITULO'] = df['TITULO'].astype(str) + '  -'
df['EDIT'] = df['EDIT'].str.slice(stop=-3)
df['EDIT'] = df['EDIT'].str.lstrip('0')
df.replace({'BNAHU080': 'HOL', 'BNOBR080': 'B70', 'COOBR090': 'B90', 'BNOBR090': 'B90', 'BNILM115': 'E115',
            'COAHU080': 'HOL', 'TAILU270': 'ESM300', 'ENCBIN': 'RÚST.', 'ENCACA': 'CABA.', 'LAMMAT': 'MAT',
            'LAMBTE': 'BTE', 'COILM090': 'MAT90'}, regex=True, inplace=True)
df['EDIT'] = df.pop('EDIT')
df['PWSH'] = df['COD_PUB']
df['PWSH'] = df['PWSH'].str.lstrip('0') + '*,'
new_file_name = "_1a1_.xlsx"
df.to_excel(new_file_name, index=False)
print(f"File saved as {new_file_name}")
wb = openpyxl.load_workbook('_1a1_.xlsx')
ws = wb.active
fill = PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'MAT' in str(cell.value):
            cell.fill = fill
for row in ws.iter_rows():
    for cell in row:
        if 'E115' in str(cell.value):
            cell.fill = PatternFill(start_color='A569BD', end_color='A569BD', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'CL' in str(cell.value):
            cell.fill = PatternFill(start_color='EC7063', end_color='EC7063', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if cell.value == None:
            cell.fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for row in ws.iter_rows(min_row=2):
    if row[3].value != None and isinstance(row[3].value, (int, float)) and float(row[3].value) > 1:
        row[3].fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for cell in ws[1]:
    cell.fill = PatternFill(start_color='AEB6BF', end_color='AEB6BF', fill_type='solid')
wb.save('_1a1_.xlsx')
wb = openpyxl.load_workbook('_1a1_.xlsx')
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.1
    ws.column_dimensions[column].width = adjusted_width
wb.save('_1a1_.xlsx')
fil1 = pd.read_excel('_1a1_.xlsx', usecols=[1])
names = fil1.values.flatten().tolist()
dir_path = "./"
files = os.listdir(dir_path)
###
ws.append([])
ws.cell(row=ws.max_row + 2, column=1, value="CONT.")
ws.cell(row=ws.max_row, column=2, value="CANT.")
ws.cell(row=ws.max_row, column=3, value="PLIEGO")
ws.cell(row=ws.max_row, column=4, value="CORTE")
for row in ws.iter_rows():
    for cell in row:
        if cell.value in ["CONT.", "CANT.", "PLIEGO", "CORTE"]:
            cell.fill = PatternFill(start_color='AEB6BF', end_color='AEB6BF', fill_type='solid')
wb.save('_1a1_.xlsx')
import math
wb = openpyxl.load_workbook('_1a1_.xlsx')
ws = wb.active
B70X = 0.0
B70 = 0.0
B90X = 0.0
B90 = 0.0
HOLX = 0.0
HOL = 0.0
for i in range(1, ws.max_row + 1):
    if ws.cell(row=i, column=8).value == "B70" and ws.cell(row=i, column=6).value >= "170x240":
        B70X += ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value
    elif ws.cell(row=i, column=8).value == "HOL" and ws.cell(row=i, column=6).value >= "170x240":
        HOLX += ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value
    elif ws.cell(row=i, column=8).value == "B90" and ws.cell(row=i, column=6).value >= "170x240":
        B90X += ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value
    elif ws.cell(row=i, column=8).value == "B70" and ws.cell(row=i, column=6).value < "170x240":
        B70 += ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value
    elif ws.cell(row=i, column=8).value == "HOL" and ws.cell(row=i, column=6).value < "170x240":
        HOL += ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value
    elif ws.cell(row=i, column=8).value == "B90" and ws.cell(row=i, column=6).value < "170x240":
        B90 += ws.cell(row=i, column=5).value * ws.cell(row=i, column=4).value
B70X = math.ceil(B70X / 16)
HOLX = math.ceil(HOLX / 16)
B90X = math.ceil(B90X / 16)
B70 = math.ceil(B70 / 32)
HOL = math.ceil(HOL / 32)
B90 = math.ceil(B90 / 32)
max_row = ws.max_row
ws.cell(row=max_row + 1, column=1, value="B70")
ws.cell(row=max_row + 2, column=1, value="B70")
ws.cell(row=max_row + 3, column=1, value="B90")
ws.cell(row=max_row + 4, column=1, value="B90")
ws.cell(row=max_row + 5, column=1, value="HOL")
ws.cell(row=max_row + 6, column=1, value="HOL")
ws.cell(row=max_row + 1, column=2, value=B70X)
ws.cell(row=max_row + 2, column=2, value=B70)
ws.cell(row=max_row + 3, column=2, value=B90X)
ws.cell(row=max_row + 4, column=2, value=B90)
ws.cell(row=max_row + 5, column=2, value=HOLX)
ws.cell(row=max_row + 6, column=2, value=HOL)
ws.cell(row=max_row + 1, column=3, value="70X100")
ws.cell(row=max_row + 2, column=3, value="70X100")
ws.cell(row=max_row + 3, column=3, value="70X100")
ws.cell(row=max_row + 4, column=3, value="70X100")
ws.cell(row=max_row + 5, column=3, value="70X100")
ws.cell(row=max_row + 6, column=3, value="70X100")
ws.cell(row=max_row + 1, column=4, value="35X50")
ws.cell(row=max_row + 2, column=4, value="25X35")
ws.cell(row=max_row + 3, column=4, value="35X50")
ws.cell(row=max_row + 4, column=4, value="25X35")
ws.cell(row=max_row + 5, column=4, value="35X50")
ws.cell(row=max_row + 6, column=4, value="25X35")
wb.save('_1a1_.xlsx')
