###Roygbiv
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import os

df = pd.read_excel('IgetBored.xlsx')
df.drop(df.columns[[1, 2, 3, 4, 8, 17, 18, 19]], axis=1, inplace=True)
df.iloc[:, 0] = df.iloc[:, 0].str[5:]
df.iloc[:, 1] = df.iloc[:, 1].str[5:]
df.iloc[:, 3] = df.iloc[:, 3].str[6:]
df['COD_PUB'] = df['COD_PUB'].str.lstrip('0') + '_'
df['TITULO'] = df['TITULO'].str.slice(stop=30)
df['TITULO'] = df['TITULO'].astype(str) + '  -'
df['EDIT'] = df['EDIT'].str.slice(stop=-3)
df.replace({'BNAHU080': 'HOL.', 'BNOBR080': 'B70', 'COOBR090': 'B90', 'BNOBR090': 'B90', 'BNILM115': 'E115',
            'COAHU080': 'HOL.', 'TAILU270': 'ESM300', 'ENCBIN': 'RÚST.', 'ENCACA': 'CABA.', 'LAMMAT': 'MAT',
            'LAMBTE': 'BTE'}, regex=True, inplace=True)
df['EDIT'] = df.pop('EDIT')
df['PWSH'] = df['COD_PUB']
df['PWSH'] = df['PWSH'].str.lstrip('0') + '*,'
new_file_name = "_1a1BMG_.xlsx"
df.to_excel(new_file_name, index=False)
print(f"File saved as {new_file_name}")
wb = openpyxl.load_workbook('_1a1BMG_.xlsx')
ws = wb.active
fill = PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'MAT' in str(cell.value):
            cell.fill = fill
for row in ws.iter_rows():
    for cell in row:
        if 'E115' in str(cell.value):
            cell.fill = PatternFill(start_color='A569BD', end_color='A569BD', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'CL' in str(cell.value):
            cell.fill = PatternFill(start_color='EC7063', end_color='EC7063', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if cell.value == None:
            cell.fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for row in ws.iter_rows(min_row=2):
    if row[3].value != None and isinstance(row[3].value, (int, float)) and float(row[3].value) > 1:
        row[3].fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for cell in ws[1]:
    cell.fill = PatternFill(start_color='AEB6BF', end_color='AEB6BF', fill_type='solid')
wb.save('_1a1BMG_.xlsx')
wb = openpyxl.load_workbook('_1a1BMG_.xlsx')
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.1
    ws.column_dimensions[column].width = adjusted_width
wb.save('_1a1BMG_.xlsx')
fil1 = pd.read_excel('_1a1BMG_.xlsx', usecols=[1])
names = fil1.values.flatten().tolist()
dir_path = "./"
files = os.listdir(dir_path)
wb = openpyxl.load_workbook('_1a1BMG_.xlsx')
ws = wb.active
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
for i, name in enumerate(names):
    found = False
    for file in files:
        if name in file and file.startswith(name.split('_')[0]):
            found = True
            break
    if not found:
        print(f'{name} not found in directory')
        ws.cell(row=i+2, column=2).fill = red_fill
wb.save('_1a1BMG_.xlsx')
###
###
###
df = pd.read_excel('_1a1BMG_.xlsx')
df['ANCHOALTO'] = df['ANCHOALTO'].str.replace('x', '').astype(int)
df['result'] = df.apply(lambda row: row['CANT'] * row['PAG'] / 16 if int(row['ANCHOALTO']) <= 170240 else row['CANT'] * row['PAG'] / 32, axis=1)
results = {'B70': [], 'B90': [], 'HOL': [], 'ESM115': []}
for index, row in df.iterrows():
    if row['CONT'] == 'B70':
        results['B70'].append(row['CANT'] * row['PAG'] / 32 if row['ANCHOALTO'] < 170240 else row['CANT'] * row['PAG'] / 16)
    elif row['CONT'] == 'B90':
        results['B90'].append(row['CANT'] * row['PAG'] / 32 if row['ANCHOALTO'] < 170240 else row['CANT'] * row['PAG'] / 16)
    elif row['CONT'] == 'HOL':
        results['HOL'].append(row['CANT'] * row['PAG'] / 32 if row['ANCHOALTO'] < 170240 else row['CANT'] * row['PAG'] / 16)
df.loc[df.index[-1] + 1] = ['' for _ in range(len(df.columns))]
df.iloc[-1, 3] = sum(results['B70'])
df.iloc[-1, 4] = sum(results['B90'])
df.iloc[-1, 5] = sum(results['HOL'])
