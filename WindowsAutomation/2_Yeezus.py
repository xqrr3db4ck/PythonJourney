import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import os

def remove_files_with_words(directory, words):
    files = os.listdir(directory)
    for file in files:
        if any(word in file for word in words):
            os.remove(os.path.join(directory, file))
def remove_duplicate_files(directory):
    files = os.listdir(directory)
    tapa_files = {}
    contenido_files = {}
    for file in files:
        if "TAPA" in file:
            name = file.split("_")[0]
            if name not in tapa_files:
                tapa_files[name] = []
            tapa_files[name].append(file)
        elif "CONTENIDO" in file:
            name = file.split("_")[0]
            if name not in contenido_files:
                contenido_files[name] = []
            contenido_files[name].append(file)
    for files_dict in [tapa_files, contenido_files]:
        for name, files in files_dict.items():
            if len(files) > 1:
                files.sort()
                for file in files[:-1]:
                    os.remove(os.path.join(directory, file))
directory = dest_dir
words = ["MUESTRA", "IMAGEN", "THECAKEISALIE"]
remove_files_with_words(directory, words)
directory = dest_dir
remove_duplicate_files(directory)

fil2 = pd.read_excel('_1a1BMG_.xlsx', usecols=[1])
names = fil2.values.flatten().tolist()
dir_path = dest_dir
files = os.listdir(dir_path)
wb = openpyxl.load_workbook('_1a1BMG_.xlsx')
ws = wb.active
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
for i, name in enumerate(names):
    found = False
    for file in files:
        if name in file and file.startswith(name.split('_')[0]):
            found = True
            break
    if not found:
        print(f'{name} not found in directory')
        ws.cell(row=i+2, column=2).fill = red_fill
wb.save('_1a1BMG_.xlsx')
