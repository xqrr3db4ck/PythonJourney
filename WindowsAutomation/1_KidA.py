import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

df = pd.read_excel('IgetBored.xlsx')
df.drop(df.columns[[1, 2, 3, 4, 8, 17, 18, 19]], axis=1, inplace=True)
df.iloc[:, 0] = df.iloc[:, 0].str[5:]
df.iloc[:, 1] = df.iloc[:, 1].str[5:]
df.iloc[:, 3] = df.iloc[:, 3].str[6:]
df['COD_PUB'] = df['COD_PUB'].str.lstrip('0') + '_'
df['TITULO'] = df['TITULO'].str.slice(stop=30)
df['TITULO'] = df['TITULO'].astype(str) + '  -'
df['EDIT'] = df['EDIT'].str.slice(stop=-3)
df.replace({'BNAHU080': 'HOL.', 'BNOBR080': 'B70', 'COOBR090': 'B90', 'BNOBR090': 'B90', 'BNILM115': 'E115',
            'COAHU080': 'HOL.', 'TAILU270': 'ESM300', 'ENCBIN': 'RÚST.', 'ENCACA': 'CABA.', 'LAMMAT': 'MAT',
            'LAMBTE': 'BTE'}, regex=True, inplace=True)
df['EDIT'] = df.pop('EDIT')
df['PWSH'] = df['COD_PUB']
df['PWSH'] = df['PWSH'].str.lstrip('0') + '*,'
new_file_name = "_1a1BMG_.xlsx"
df.to_excel(new_file_name, index=False)
print(f"File saved as {new_file_name}")
wb = openpyxl.load_workbook('_1a1BMG_.xlsx')
ws = wb.active
fill = PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'MAT' in str(cell.value):
            cell.fill = fill
for row in ws.iter_rows():
    for cell in row:
        if 'E115' in str(cell.value):
            cell.fill = PatternFill(start_color='A569BD', end_color='A569BD', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'CL' in str(cell.value):
            cell.fill = PatternFill(start_color='EC7063', end_color='EC7063', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if cell.value == None:
            cell.fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for row in ws.iter_rows(min_row=2):
    if row[3].value != None and isinstance(row[3].value, (int, float)) and float(row[3].value) > 1:
        row[3].fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for cell in ws[1]:
    cell.fill = PatternFill(start_color='AEB6BF', end_color='AEB6BF', fill_type='solid')
wb.save('_1a1BMG_.xlsx')
wb = openpyxl.load_workbook('_1a1BMG_.xlsx')
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.1
    ws.column_dimensions[column].width = adjusted_width
wb.save('_1a1BMG_.xlsx')

# Second_Part
dir11 = pd.read_excel('_1a1BMG_.xlsx', usecols=[11])
fil1 = pd.read_excel('_1a1BMG_.xlsx', usecols=[1])
print(dir11)
print(fil1)

dest_dir = input("Enter path: ")
print(dest_dir + ' Will be the destiny folder')
root_dir = 'O:'
subdirs = [x[0] for x in os.walk(root_dir)]
for subdir in subdirs:
    for val in dir12.values:
        if str(val[0]) in subdir:
            for filename in os.listdir(subdir):
                for fileval in fil2.values:
                    if str(fileval[0]) in filename and filename.endswith('.pdf'):
                        source_file = os.path.join(subdir, filename)
                        dest_file = os.path.join(dest_dir, filename)
                        if not os.path.exists(dest_file):
                            shutil.copy(source_file, dest_file)
                        else:
                            print(f"{filename} already exists in {dest_dir}, skipping...")
os.chdir(dest_dir)
def remove_files_with_words(directory, words):
    files = os.listdir(directory)
    for file in files:
        if any(word in file for word in words):
            os.remove(os.path.join(directory, file))
def remove_duplicate_files(directory):
    files = os.listdir(directory)
    tapa_files = {}
    contenido_files = {}
    for file in files:
        if "TAPA" in file:
            name = file.split("_")[0]
            if name not in tapa_files:
                tapa_files[name] = []
            tapa_files[name].append(file)
        elif "CONTENIDO" in file:
            name = file.split("_")[0]
            if name not in contenido_files:
                contenido_files[name] = []
            contenido_files[name].append(file)
    for files_dict in [tapa_files, contenido_files]:
        for name, files in files_dict.items():
            if len(files) > 1:
                files.sort()
                for file in files[:-1]:
                    os.remove(os.path.join(directory, file))
directory = dest_dir
words = ["MUESTRA", "IMAGEN", "THECAKEISALIE"]
remove_files_with_words(directory, words)
directory = dest_dir
remove_duplicate_files(directory)

fil2 = pd.read_excel('_1a1BMG_.xlsx', usecols=[2])
names = fil2.values.flatten().tolist()
dir_path = dest_dir
files = os.listdir(dir_path)
wb = load_workbook('_1a1BMG_.xlsx')
ws = wb.active
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
for i, name in enumerate(names):
    found = False
    for file in files:
        if name in file and file.startswith(name.split('_')[0]):
            found = True
            break
    if not found:
        print(f'{name} not found in directory')
        ws.cell(row=i+2, column=3).fill = red_fill
wb.save('_1a1BMG_.xlsx')

os.makedirs(os.path.join(dest_dir, "CONTENIDOS"))
os.makedirs(os.path.join(dest_dir, "MONTAJES"))
os.makedirs(os.path.join(dest_dir, "TAPAS"))
os.makedirs(os.path.join(dest_dir, "TAPAS", "BRILLANTE"))
os.makedirs(os.path.join(dest_dir, "TAPAS", "MATE"))
source_dir = dest_dir
contents_dir = 'CONTENIDOS'
impositions_dir = 'MONTAJES'
covers_dir = 'TAPAS'
for filename in os.listdir(source_dir):
    if 'tapa' in filename.lower():
        shutil.move(os.path.join(source_dir, filename), covers_dir)
    elif 'contenido' in filename.lower():
        shutil.copy(os.path.join(source_dir, filename), impositions_dir)
        shutil.move(os.path.join(source_dir, filename), contents_dir)
