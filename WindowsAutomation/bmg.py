import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from math import ceil
import os

df = pd.read_excel('bmg.xlsx')
df.drop(df.columns[[1, 2, 4, 8, 9, 10, 11, 15, 20, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37]], axis=1, inplace=True)
if (df.iloc[:, 10] == 0).all() or (df.iloc[:, 11] == "0").all():
    df.drop(df.columns[[8, 10, 11]], axis=1, inplace=True)
df['COD_EDIT'] = df.pop('COD_EDIT')
df.iloc[:, 0] = df.iloc[:, 0].str[5:]
df.iloc[:, 1] = df.iloc[:, 1].str[5:]
df['COD_PUB'] = df['COD_PUB'].str.lstrip('0') + '_'
df['TITULO'] = df['TITULO'].str.slice(stop=30)
df['TITULO'] = df['TITULO'].astype(str) + '  -'
df.replace({'BNAHU080': 'HOL.', 'BNOBR080': 'B70', 'COOBR090': 'B90', 'BNOBR090': 'B90', 'BNILM115': 'E115',
            'COAHU080': 'HOL.', 'TAILU270': 'ESM300', 'ENCBIN': 'RÚST.', 'ENCACA': 'CABA.', 'LAMMAT': 'MAT',
            'LAMBTE': 'BTE', 'COILM090': 'MAT90'}, regex=True, inplace=True)
df['PWSH'] = df['COD_PUB']
df['PWSH'] = df['PWSH'].str.lstrip('0') + '*,'
new_file_name = "_bmg_.xlsx"
df.to_excel(new_file_name, index=False)
print(f"File saved as {new_file_name}")
###
wb = openpyxl.load_workbook('_bmg_.xlsx')
ws = wb.active
for cell in ws[1]:
    value = str(cell.value)
    if value == 'ENCUADERNACION':
        cell.value = 'ENC.'
    elif value == 'LAMINADO':
        cell.value = 'LAM.'
    elif value == 'PAG_TOTALES':
        cell.value = 'PAG'
    elif value == 'PAPEL_TAPA':
        cell.value = 'TAPA'
    elif value == 'PAPEL_BN':
        cell.value = 'CONT.'
    elif value == 'CANTIDAD':
        cell.value = 'CANT.'
    elif value == 'COD_EDIT':
        cell.value = 'EDIT.'
    wb.save('_bmg_.xlsx')
###
wb = openpyxl.load_workbook('_bmg_.xlsx')
ws = wb.active
fill = PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'MAT' in str(cell.value):
            cell.fill = fill
for row in ws.iter_rows():
    for cell in row:
        if 'E115' in str(cell.value):
            cell.fill = PatternFill(start_color='A569BD', end_color='A569BD', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if 'CL' in str(cell.value):
            cell.fill = PatternFill(start_color='EC7063', end_color='EC7063', fill_type='solid')
for row in ws.iter_rows():
    for cell in row:
        if cell.value == None:
            cell.fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for row in ws.iter_rows(min_row=2):
    if row[3].value != None and isinstance(row[3].value, (int, float)) and float(row[3].value) > 1:
        row[3].fill = PatternFill(start_color='F7DC6F', end_color='F7DC6F', fill_type='solid')
for cell in ws[1]:
    cell.fill = PatternFill(start_color='AEB6BF', end_color='AEB6BF', fill_type='solid')
wb.save('_bmg_.xlsx')
wb = openpyxl.load_workbook('_bmg_.xlsx')
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.1
    ws.column_dimensions[column].width = adjusted_width
wb.save('_bmg_.xlsx')
fil1 = pd.read_excel('_bmg_.xlsx', usecols=[1])
names = fil1.values.flatten().tolist()
dir_path = "./"
files = os.listdir(dir_path)
#####
ws.append([])
ws.cell(row=ws.max_row + 2, column=1, value="CONT.")
ws.cell(row=ws.max_row, column=2, value="CANT.")
ws.cell(row=ws.max_row, column=3, value="PLIEGO")
ws.cell(row=ws.max_row, column=4, value="CORTE")
for row in ws.iter_rows():
    for cell in row:
        if cell.value in ["CONT.", "CANT.", "PLIEGO", "CORTE"]:
            cell.fill = PatternFill(start_color='AEB6BF', end_color='AEB6BF', fill_type='solid')
wb.save('_bmg_.xlsx')
