import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

fil2 = pd.read_excel('PedidosBMG1a1.xlsx', usecols=[2])
names = fil2.values.flatten().tolist()
dir_path = 'D:/pythonProject/Lab'
files = os.listdir(dir_path)
wb = load_workbook('PedidosBMG1a1.xlsx')
ws = wb.active
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
for i, name in enumerate(names):
    found = False
    for file in files:
        if name in file and file.startswith(name.split('_')[0]):
            found = True
            break
    if not found:
        print(f'{name} not found in directory')
        ws.cell(row=i+2, column=3).fill = red_fill
wb.save('PedidosBMG1a1.xlsx')

'''
import pandas as pd
import os

fil2 = pd.read_excel('PedidosBMG1a1.xlsx', usecols=[2])
names = fil2.values.flatten().tolist()
dir_path = 'D:/pythonProject/Lab'
files = os.listdir(dir_path)
for name in names:
    found = False
    for file in files:
        if name in file and file.startswith(name.split('_')[0]):
            found = True
            break
    if not found:
        print(f'{name} not found in directory')
'''
